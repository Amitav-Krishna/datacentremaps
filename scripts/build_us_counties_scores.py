#!/usr/bin/env python3
"""
Build scored US county boundaries from raw source datasets.

Outputs:
- data/derived/us_counties_scored.geojson
- data/derived/us_counties_scored_manifest.json
- data/derived/us_counties_scored.geojson.gz (optional)
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


FIPS_TO_ABBR = {
    "01": "AL",
    "02": "AK",
    "04": "AZ",
    "05": "AR",
    "06": "CA",
    "08": "CO",
    "09": "CT",
    "10": "DE",
    "11": "DC",
    "12": "FL",
    "13": "GA",
    "15": "HI",
    "16": "ID",
    "17": "IL",
    "18": "IN",
    "19": "IA",
    "20": "KS",
    "21": "KY",
    "22": "LA",
    "23": "ME",
    "24": "MD",
    "25": "MA",
    "26": "MI",
    "27": "MN",
    "28": "MS",
    "29": "MO",
    "30": "MT",
    "31": "NE",
    "32": "NV",
    "33": "NH",
    "34": "NJ",
    "35": "NM",
    "36": "NY",
    "37": "NC",
    "38": "ND",
    "39": "OH",
    "40": "OK",
    "41": "OR",
    "42": "PA",
    "44": "RI",
    "45": "SC",
    "46": "SD",
    "47": "TN",
    "48": "TX",
    "49": "UT",
    "50": "VT",
    "51": "VA",
    "53": "WA",
    "54": "WV",
    "55": "WI",
    "56": "WY",
    "72": "PR",
}

STATE_NAME_TO_ABBR = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
}


def normalize_inverse(value: float, vmin: float, vmax: float) -> float:
    if vmax == vmin:
        return 50.0
    clamped = max(vmin, min(vmax, value))
    return 100.0 * (1.0 - (clamped - vmin) / (vmax - vmin))


def normalize(value: float, vmin: float, vmax: float) -> float:
    if vmax == vmin:
        return 50.0
    clamped = max(vmin, min(vmax, value))
    return 100.0 * (clamped - vmin) / (vmax - vmin)


def bbox_for_features(features: list[dict]) -> list[float] | None:
    min_x = min_y = float("inf")
    max_x = max_y = float("-inf")

    def scan(coords):
        nonlocal min_x, min_y, max_x, max_y
        if isinstance(coords, (list, tuple)):
            if coords and isinstance(coords[0], (int, float)) and len(coords) >= 2:
                x, y = float(coords[0]), float(coords[1])
                min_x = min(min_x, x)
                min_y = min(min_y, y)
                max_x = max(max_x, x)
                max_y = max(max_y, y)
            else:
                for item in coords:
                    scan(item)

    for feature in features:
        geometry = feature.get("geometry") or {}
        scan(geometry.get("coordinates"))

    if min_x == float("inf"):
        return None
    return [min_x, min_y, max_x, max_y]


def load_state_power_prices(path: Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    years = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == "Total Electric Industry":
            years.add(row[0])
    latest_year = max(years)

    state_power_price: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == latest_year and row[2] == "Total Electric Industry" and row[1] != "US":
            state_abbr = row[1]
            price = row[5]
            if isinstance(price, (int, float)):
                state_power_price[state_abbr] = float(price)
    print(f"Loaded electricity prices for {len(state_power_price)} states (year {latest_year})")
    return state_power_price


def load_county_home_values(path: Path) -> dict[str, int]:
    with path.open("r", encoding="utf-8") as f:
        census_data = json.load(f)

    county_home_value: dict[str, int] = {}
    for row in census_data[1:]:
        _, value, state_fips, county_fips = row
        fips = state_fips + county_fips
        if value and value != "null":
            county_home_value[fips] = int(value)
    print(f"Loaded home values for {len(county_home_value)} counties")
    return county_home_value


def load_county_permits_per_capita(permits_path: Path, population_path: Path) -> dict[str, float]:
    county_permits: dict[str, int] = {}
    with permits_path.open("r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            if i < 2 or not line.strip():
                continue
            parts = line.split(",")
            if len(parts) < 8:
                continue
            state_fips = parts[1].strip().zfill(2)
            county_fips = parts[2].strip().zfill(3)
            fips = state_fips + county_fips
            try:
                total_units = int(parts[7].strip())
                county_permits[fips] = county_permits.get(fips, 0) + total_units
            except (ValueError, IndexError):
                continue

    county_pop: dict[str, int] = {}
    with population_path.open("r", encoding="utf-8") as f:
        pop_data = json.load(f)
    for row in pop_data[1:]:
        _, pop, state_fips, county_fips = row
        fips = state_fips + county_fips
        if pop and pop != "null":
            county_pop[fips] = int(pop)

    county_permits_pc: dict[str, float] = {}
    for fips, permits in county_permits.items():
        pop = county_pop.get(fips)
        if pop and pop > 0:
            county_permits_pc[fips] = permits / pop * 1000.0

    print(
        f"Loaded permits for {len(county_permits)} counties, "
        f"per-capita for {len(county_permits_pc)}"
    )
    return county_permits_pc


def load_state_regulatory_scores(path: Path) -> dict[str, float]:
    wb_cato = openpyxl.load_workbook(path, data_only=True)
    ws_cato = wb_cato["Overall"]
    state_reg_score: dict[str, tuple[int, float]] = {}
    for row_idx in range(2, ws_cato.max_row + 1):
        state_name = ws_cato.cell(row_idx, 1).value
        year = ws_cato.cell(row_idx, 2).value
        reg_score = ws_cato.cell(row_idx, 5).value
        if state_name and year and reg_score is not None:
            abbr = STATE_NAME_TO_ABBR.get(state_name)
            if abbr:
                current = state_reg_score.get(abbr)
                if current is None or year > current[0]:
                    state_reg_score[abbr] = (year, float(reg_score))
    state_reg = {abbr: year_score[1] for abbr, year_score in state_reg_score.items()}
    print(f"Loaded regulatory scores for {len(state_reg)} states")
    return state_reg


def load_county_broadband_tier(path: Path) -> dict[str, int]:
    county_broadband: dict[str, tuple[tuple[int, int], int]] = {}
    with path.open("r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for row in reader:
            fips = row["FIPS"]
            key = (int(row["Year"]), int(row["Month"]))
            tier = int(row["Tier_1"])
            if fips not in county_broadband or key > county_broadband[fips][0]:
                county_broadband[fips] = (key, tier)
    county_bb = {fips: value[1] for fips, value in county_broadband.items()}
    print(f"Loaded broadband tiers for {len(county_bb)} counties")
    return county_bb


def add_scores_to_geojson(
    geojson: dict,
    state_power_price: dict[str, float],
    county_home_value: dict[str, int],
    county_permits_pc: dict[str, float],
    state_reg: dict[str, float],
    county_bb: dict[str, int],
) -> dict:
    power_values = list(state_power_price.values())
    power_min, power_max = min(power_values), max(power_values)

    home_values = [v for v in county_home_value.values() if v > 0]
    home_values_sorted = sorted(home_values)
    home_min = min(home_values)
    home_p95 = home_values_sorted[int(len(home_values_sorted) * 0.95)]

    permits_vals = [v for v in county_permits_pc.values() if v > 0]
    permits_sorted = sorted(permits_vals)
    permits_min = min(permits_vals)
    permits_p95 = permits_sorted[int(len(permits_sorted) * 0.95)]

    reg_vals = list(state_reg.values())
    reg_min, reg_max = min(reg_vals), max(reg_vals)

    print(f"Power price range: {power_min} - {power_max} cents/kWh")
    print(f"Home value range: ${int(home_min):,} - ${int(home_p95):,} (p95)")
    print(f"Permits/1k pop range: {permits_min:.2f} - {permits_p95:.2f} (p95)")
    print(f"Regulatory score range: {reg_min:.3f} - {reg_max:.3f}")
    print("Broadband tier range: 0 - 5")

    missing_power = 0
    missing_land = 0
    missing_permits = 0
    missing_reg = 0
    missing_bb = 0
    scored = 0

    for feature in geojson.get("features", []):
        props = feature.get("properties") or {}
        state_fips = str(props.get("STATE", "")).zfill(2)
        county_fips = str(props.get("COUNTY", "")).zfill(3)
        fips = state_fips + county_fips
        state_abbr = FIPS_TO_ABBR.get(state_fips)

        power_score = None
        if state_abbr and state_abbr in state_power_price:
            power_score = normalize_inverse(state_power_price[state_abbr], power_min, power_max)
        else:
            missing_power += 1

        land_score = None
        if fips in county_home_value and county_home_value[fips] > 0:
            land_score = normalize_inverse(county_home_value[fips], home_min, home_p95)
        else:
            missing_land += 1

        permit_score = None
        if fips in county_permits_pc:
            permit_score = normalize(county_permits_pc[fips], permits_min, permits_p95)
        else:
            missing_permits += 1

        reg_score_val = None
        if state_abbr and state_abbr in state_reg:
            reg_score_val = normalize(state_reg[state_abbr], reg_min, reg_max)
        else:
            missing_reg += 1

        bb_score = None
        if fips in county_bb:
            bb_score = normalize(county_bb[fips], 0, 5)
        else:
            missing_bb += 1

        props["s_power"] = round(power_score, 1) if power_score is not None else None
        props["s_land"] = round(land_score, 1) if land_score is not None else None
        props["s_permits"] = round(permit_score, 1) if permit_score is not None else None
        props["s_reg"] = round(reg_score_val, 1) if reg_score_val is not None else None
        props["s_broadband"] = round(bb_score, 1) if bb_score is not None else None

        props["power_price"] = state_power_price.get(state_abbr) if state_abbr else None
        props["home_value"] = county_home_value.get(fips)
        props["permits_pc"] = round(county_permits_pc[fips], 2) if fips in county_permits_pc else None
        props["reg_freedom"] = (
            round(state_reg[state_abbr], 3) if (state_abbr and state_abbr in state_reg) else None
        )
        props["broadband_tier"] = county_bb.get(fips)

        layer_scores = [
            score
            for score in [power_score, land_score, permit_score, reg_score_val, bb_score]
            if score is not None
        ]
        props["score"] = round(sum(layer_scores) / len(layer_scores), 1) if layer_scores else 50.0
        scored += 1

    print(f"Scored {scored} counties")
    print(
        "Missing:"
        f" power={missing_power}, land={missing_land}, permits={missing_permits},"
        f" reg={missing_reg}, broadband={missing_bb}"
    )
    return geojson


def write_outputs(output_geojson_path: Path, output_manifest_path: Path, geojson: dict, write_gzip: bool) -> None:
    output_geojson_path.parent.mkdir(parents=True, exist_ok=True)
    output_manifest_path.parent.mkdir(parents=True, exist_ok=True)

    payload = json.dumps(geojson, separators=(",", ":"), ensure_ascii=False)
    output_geojson_path.write_text(payload, encoding="utf-8")
    payload_bytes = payload.encode("utf-8")
    payload_sha = hashlib.sha256(payload_bytes).hexdigest()

    features = geojson.get("features", [])
    manifest = {
        "artifact": str(output_geojson_path).replace("\\", "/"),
        "feature_count": len(features),
        "bbox": bbox_for_features(features),
        "sha256": payload_sha,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    output_manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    if write_gzip:
        gz_path = output_geojson_path.with_suffix(output_geojson_path.suffix + ".gz")
        with gzip.open(gz_path, "wb", compresslevel=6) as gz_file:
            gz_file.write(payload_bytes)
        print(f"Wrote {gz_path}")

    print(f"Wrote {output_geojson_path}")
    print(f"Wrote {output_manifest_path}")
    print(f"feature_count={manifest['feature_count']}")
    print(f"sha256={manifest['sha256']}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build scored US county boundaries from raw data.")
    parser.add_argument("--counties-geojson", default="data/raw/us_counties.geojson")
    parser.add_argument("--eia-xlsx", default="data/raw/eia_avgprice.xlsx")
    parser.add_argument("--home-values-json", default="data/raw/census_home_values.json")
    parser.add_argument("--permits-txt", default="data/raw/permitting/bps_annual.txt")
    parser.add_argument("--population-json", default="data/raw/permitting/county_population.json")
    parser.add_argument("--regulatory-xlsx", default="data/raw/freedominthe50states.xlsx")
    parser.add_argument(
        "--broadband-csv",
        default="data/raw/fibre/county_tiers_201406_202406/county_tiers_201406_202406.csv",
    )
    parser.add_argument("--output-geojson", default="data/derived/us_counties_scored.geojson")
    parser.add_argument("--output-manifest", default="data/derived/us_counties_scored_manifest.json")
    parser.add_argument("--no-gzip", action="store_true", help="Do not write .geojson.gz output.")
    return parser.parse_args()


def require_paths(paths: list[Path]) -> None:
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        joined = "\n  - ".join(missing)
        raise FileNotFoundError(f"Missing input files:\n  - {joined}")


def main() -> None:
    args = parse_args()

    counties_geojson = Path(args.counties_geojson)
    eia_xlsx = Path(args.eia_xlsx)
    home_values_json = Path(args.home_values_json)
    permits_txt = Path(args.permits_txt)
    population_json = Path(args.population_json)
    regulatory_xlsx = Path(args.regulatory_xlsx)
    broadband_csv = Path(args.broadband_csv)
    output_geojson = Path(args.output_geojson)
    output_manifest = Path(args.output_manifest)

    require_paths(
        [
            counties_geojson,
            eia_xlsx,
            home_values_json,
            permits_txt,
            population_json,
            regulatory_xlsx,
            broadband_csv,
        ]
    )

    with counties_geojson.open("r", encoding="utf-8") as f:
        geojson = json.load(f)

    scored_geojson = add_scores_to_geojson(
        geojson=geojson,
        state_power_price=load_state_power_prices(eia_xlsx),
        county_home_value=load_county_home_values(home_values_json),
        county_permits_pc=load_county_permits_per_capita(permits_txt, population_json),
        state_reg=load_state_regulatory_scores(regulatory_xlsx),
        county_bb=load_county_broadband_tier(broadband_csv),
    )
    write_outputs(
        output_geojson_path=output_geojson,
        output_manifest_path=output_manifest,
        geojson=scored_geojson,
        write_gzip=not args.no_gzip,
    )


if __name__ == "__main__":
    main()

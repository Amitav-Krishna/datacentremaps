"""
Pre-process real data into county scores and bake into the GeoJSON.
Factors:
  1. Electricity price (state-level industrial rate, EIA) — lower is better
  2. Land cost (county-level median home value, Census ACS) — lower is better
  3. Permitting ease (county-level building permits per capita, Census BPS) — higher is better
  4. Regulatory freedom (state-level, Cato Freedom in the 50 States) — higher is better

Output: counties_scored.geojson with a 'score' property per feature.
"""
import json
import os
import csv
import openpyxl

# --- 1. Load electricity prices (state-level, industrial) ---
wb = openpyxl.load_workbook("eia_avgprice.xlsx")
ws = wb.active

years = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2] == "Total Electric Industry":
        years.add(row[0])
latest_year = max(years)

state_power_price = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] == latest_year and row[2] == "Total Electric Industry" and row[1] != "US":
        state_abbr = row[1]
        price = row[5]  # Industrial column
        if isinstance(price, (int, float)):
            state_power_price[state_abbr] = price

print(f"Loaded electricity prices for {len(state_power_price)} states (year {latest_year})")

FIPS_TO_ABBR = {
    "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA",
    "08": "CO", "09": "CT", "10": "DE", "11": "DC", "12": "FL",
    "13": "GA", "15": "HI", "16": "ID", "17": "IL", "18": "IN",
    "19": "IA", "20": "KS", "21": "KY", "22": "LA", "23": "ME",
    "24": "MD", "25": "MA", "26": "MI", "27": "MN", "28": "MS",
    "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI",
    "45": "SC", "46": "SD", "47": "TN", "48": "TX", "49": "UT",
    "50": "VT", "51": "VA", "53": "WA", "54": "WV", "55": "WI",
    "56": "WY", "72": "PR",
}

# --- 2. Load land costs (county-level median home value) ---
with open("census_home_values.json") as f:
    census_data = json.load(f)

county_home_value = {}
for row in census_data[1:]:
    name, value, state_fips, county_fips = row
    fips = state_fips + county_fips
    if value and value != "null":
        county_home_value[fips] = int(value)

print(f"Loaded home values for {len(county_home_value)} counties")

# --- 3. Load building permits (county-level) ---
county_permits = {}
with open("data/permitting/bps_annual.txt") as f:
    for i, line in enumerate(f):
        if i < 2 or not line.strip():
            continue
        parts = line.split(",")
        if len(parts) < 8:
            continue
        state_fips = parts[1].strip().zfill(2)
        county_fips = parts[2].strip().zfill(3)
        fips = state_fips + county_fips
        try:
            total_units = int(parts[7].strip())  # 1-unit buildings
            county_permits[fips] = county_permits.get(fips, 0) + total_units
        except (ValueError, IndexError):
            pass

# Load population for per-capita
county_pop = {}
with open("data/permitting/county_population.json") as f:
    pop_data = json.load(f)
for row in pop_data[1:]:
    name, pop, state_fips, county_fips = row
    fips = state_fips + county_fips
    if pop and pop != "null":
        county_pop[fips] = int(pop)

# Compute permits per 1000 people
county_permits_pc = {}
for fips, permits in county_permits.items():
    if fips in county_pop and county_pop[fips] > 0:
        county_permits_pc[fips] = permits / county_pop[fips] * 1000

print(f"Loaded permits for {len(county_permits)} counties, per-capita for {len(county_permits_pc)}")

# --- 4. Load regulatory freedom (state-level, Cato) ---
STATE_NAME_TO_ABBR = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY",
}

wb_cato = openpyxl.load_workbook("freedominthe50states.xlsx", data_only=True)
ws_cato = wb_cato["Overall"]

# Get most recent regulatory score per state
state_reg_score = {}
for r in range(2, ws_cato.max_row + 1):
    state_name = ws_cato.cell(r, 1).value
    year = ws_cato.cell(r, 2).value
    reg_score = ws_cato.cell(r, 5).value  # Regulatory Policy column
    if state_name and year and reg_score is not None:
        abbr = STATE_NAME_TO_ABBR.get(state_name)
        if abbr:
            if abbr not in state_reg_score or year > state_reg_score[abbr][0]:
                state_reg_score[abbr] = (year, reg_score)

# Extract just the scores
state_reg = {k: v[1] for k, v in state_reg_score.items()}
print(f"Loaded regulatory scores for {len(state_reg)} states")

# --- 5. Load broadband tiers (county-level, FCC Form 477) ---
county_broadband = {}
with open("data/fibre/county_tiers_201406_202406/county_tiers_201406_202406.csv", encoding="latin-1") as f:
    reader = csv.DictReader(f)
    for row in reader:
        fips = row["FIPS"]
        year = int(row["Year"])
        month = int(row["Month"])
        key = (year, month)
        if fips not in county_broadband or key > county_broadband[fips][0]:
            county_broadband[fips] = (key, int(row["Tier_1"]))

county_bb = {k: v[1] for k, v in county_broadband.items()}
print(f"Loaded broadband tiers for {len(county_bb)} counties")

# --- 6. Normalize to 0-100 scores ---
power_prices = list(state_power_price.values())
power_min, power_max = min(power_prices), max(power_prices)

home_values = [v for v in county_home_value.values() if v > 0]
home_values_sorted = sorted(home_values)
home_min = min(home_values)
home_p95 = home_values_sorted[int(len(home_values_sorted) * 0.95)]

permits_vals = [v for v in county_permits_pc.values() if v > 0]
permits_sorted = sorted(permits_vals)
permits_min = min(permits_vals)
permits_p95 = permits_sorted[int(len(permits_sorted) * 0.95)]

reg_vals = list(state_reg.values())
reg_min, reg_max = min(reg_vals), max(reg_vals)

print(f"Power price range: {power_min} - {power_max} cents/kWh")
print(f"Home value range: ${home_min:,} - ${int(home_p95):,} (p95)")
print(f"Permits/1k pop range: {permits_min:.2f} - {permits_p95:.2f} (p95)")
print(f"Regulatory score range: {reg_min:.3f} - {reg_max:.3f}")
print(f"Broadband tier range: 0 - 5")


def normalize_inverse(value, vmin, vmax):
    """Lower value = higher score (0-100)."""
    if vmax == vmin:
        return 50
    clamped = max(vmin, min(vmax, value))
    return 100 * (1 - (clamped - vmin) / (vmax - vmin))


def normalize(value, vmin, vmax):
    """Higher value = higher score (0-100)."""
    if vmax == vmin:
        return 50
    clamped = max(vmin, min(vmax, value))
    return 100 * (clamped - vmin) / (vmax - vmin)


# --- 7. Load GeoJSON and compute scores ---
with open("counties.geojson") as f:
    geojson = json.load(f)

WEIGHT_POWER = 0.25
WEIGHT_LAND = 0.20
WEIGHT_PERMITS = 0.20
WEIGHT_REG = 0.15
WEIGHT_BB = 0.20

scored = 0
missing_power = 0
missing_land = 0
missing_permits = 0
missing_reg = 0
missing_bb = 0

for feature in geojson["features"]:
    props = feature["properties"]
    state_fips = props["STATE"]
    county_fips = props["COUNTY"]
    fips = state_fips + county_fips
    state_abbr = FIPS_TO_ABBR.get(state_fips)

    # Power score
    power_score = None
    if state_abbr and state_abbr in state_power_price:
        power_score = normalize_inverse(state_power_price[state_abbr], power_min, power_max)
    else:
        missing_power += 1

    # Land score
    land_score = None
    if fips in county_home_value and county_home_value[fips] > 0:
        land_score = normalize_inverse(county_home_value[fips], home_min, home_p95)
    else:
        missing_land += 1

    # Permits score (higher permits per capita = easier permitting)
    permit_score = None
    if fips in county_permits_pc:
        permit_score = normalize(county_permits_pc[fips], permits_min, permits_p95)
    else:
        missing_permits += 1

    # Regulatory freedom score (higher = less regulation = better)
    reg_score_val = None
    if state_abbr and state_abbr in state_reg:
        reg_score_val = normalize(state_reg[state_abbr], reg_min, reg_max)
    else:
        missing_reg += 1

    # Broadband tier score (0-5 scale, higher = better)
    bb_score = None
    if fips in county_bb:
        bb_score = normalize(county_bb[fips], 0, 5)
    else:
        missing_bb += 1

    # Store individual layer scores (0-100) for frontend toggle support
    props["s_power"] = round(power_score, 1) if power_score is not None else None
    props["s_land"] = round(land_score, 1) if land_score is not None else None
    props["s_permits"] = round(permit_score, 1) if permit_score is not None else None
    props["s_reg"] = round(reg_score_val, 1) if reg_score_val is not None else None
    props["s_broadband"] = round(bb_score, 1) if bb_score is not None else None

    # Raw values for hover display
    props["power_price"] = state_power_price.get(state_abbr) if state_abbr else None
    props["home_value"] = county_home_value.get(fips)
    props["permits_pc"] = round(county_permits_pc[fips], 2) if fips in county_permits_pc else None
    props["reg_freedom"] = round(state_reg[state_abbr], 3) if (state_abbr and state_abbr in state_reg) else None
    props["broadband_tier"] = county_bb.get(fips)

    # Default combined score (all layers on, equal weight)
    layer_scores = [s for s in [power_score, land_score, permit_score, reg_score_val, bb_score] if s is not None]
    props["score"] = round(sum(layer_scores) / len(layer_scores), 1) if layer_scores else 50
    scored += 1

print(f"\nScored {scored} counties")
print(f"Missing: power={missing_power}, land={missing_land}, permits={missing_permits}, reg={missing_reg}, broadband={missing_bb}")

with open("counties_scored.geojson", "w") as f:
    json.dump(geojson, f)

print(f"Written to counties_scored.geojson ({os.path.getsize('counties_scored.geojson') / 1e6:.1f} MB)")
